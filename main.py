# -*- coding: utf-8 -*-
from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.scrollview import ScrollView
from kivy.uix.label import Label
from kivy.uix.button import Button
from kivy.uix.popup import Popup
from kivy.uix.filechooser import FileChooserListView
from kivy.clock import Clock
from kivy.metrics import dp
from kivy.properties import StringProperty, ListProperty
from kivy.graphics import Color, RoundedRectangle
from kivy.core.audio import SoundLoader
from datetime import datetime, timedelta
from dataclasses import dataclass
from typing import List, Optional
import openpyxl
import re
import os


@dataclass
class Course:
    name: str
    code: str
    weeks: List[int]
    teacher: str
    start_section: int
    end_section: int
    location: str
    weekday: int


SECTION_TIMES = {
    1: (8, 0), 2: (8, 50), 3: (9, 50), 4: (10, 40), 5: (11, 30),
    6: (14, 0), 7: (14, 50), 8: (15, 50), 9: (16, 40),
    10: (18, 30), 11: (19, 20), 12: (20, 10), 13: (21, 0),
}
SECTION_DURATION = 45
WEEKDAY_NAMES = ['星期一', '星期二', '星期三', '星期四', '星期五', '星期六', '星期日']


def parse_weeks(week_str: str) -> List[int]:
    weeks = []
    if not week_str:
        return weeks
    parts = week_str.replace('，', ',').split(',')
    for part in parts:
        part = part.strip()
        if '-' in part:
            try:
                start, end = part.split('-')
                weeks.extend(range(int(start), int(end) + 1))
            except ValueError:
                pass
        else:
            try:
                weeks.append(int(part))
            except ValueError:
                pass
    return sorted(set(weeks))


def parse_course_cell(cell_value: str, weekday: int) -> List[Course]:
    if not cell_value or cell_value.strip() == '':
        return []
    courses = []
    lines = cell_value.strip().split('\n')
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        if not line:
            i += 1
            continue
        name_match = re.match(r'^(.+?)\[(.+?)\]$', line)
        if name_match:
            course_name = name_match.group(1)
            course_code = name_match.group(2)
            if i + 1 < len(lines):
                detail_line = lines[i + 1].strip()
                week_match = re.search(r'(\d+(?:-\d+)?(?:,\d+(?:-\d+)?)*)\s*周', detail_line)
                weeks = parse_weeks(week_match.group(1)) if week_match else []
                teacher_match = re.search(r'\d+周\s+(\S+)', detail_line)
                teacher = teacher_match.group(1) if teacher_match else ""
                section_match = re.search(r'第(\d+)节-第(\d+)节', detail_line)
                if section_match:
                    start_section = int(section_match.group(1))
                    end_section = int(section_match.group(2))
                else:
                    single_section_match = re.search(r'第(\d+)节', detail_line)
                    if single_section_match:
                        start_section = int(single_section_match.group(1))
                        end_section = start_section
                    else:
                        start_section = 1
                        end_section = 1
                location_match = re.search(r'第\d+节[^\s]*\s+(.+)$', detail_line)
                location = location_match.group(1).strip() if location_match else ""
                course = Course(
                    name=course_name, code=course_code, weeks=weeks, teacher=teacher,
                    start_section=start_section, end_section=end_section,
                    location=location, weekday=weekday
                )
                courses.append(course)
                i += 2
            else:
                i += 1
        else:
            i += 1
    return courses


def parse_xlsx_schedule(file_path: str) -> List[Course]:
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    all_courses = []
    for row_idx in range(4, sheet.max_row + 1):
        for col_idx in range(2, sheet.max_column + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            if cell.value:
                weekday = col_idx - 1
                courses = parse_course_cell(str(cell.value), weekday)
                all_courses.extend(courses)
    return all_courses


def get_current_week(semester_start_date: datetime) -> int:
    today = datetime.now()
    delta = today - semester_start_date
    return (delta.days // 7) + 1


def get_section_time(section: int) -> datetime:
    today = datetime.now()
    hour, minute = SECTION_TIMES.get(section, (8, 0))
    return today.replace(hour=hour, minute=minute, second=0, microsecond=0)


def get_next_course(courses: List[Course], current_week: int) -> Optional[tuple]:
    now = datetime.now()
    current_weekday = now.weekday() + 1
    current_time = now.time()
    today_courses = [c for c in courses if c.weekday == current_weekday and current_week in c.weeks]
    for course in sorted(today_courses, key=lambda c: c.start_section):
        section_start = get_section_time(course.start_section)
        section_end = get_section_time(course.end_section) + timedelta(minutes=SECTION_DURATION)
        if section_start.time() > current_time:
            minutes_until = (section_start - now).total_seconds() / 60
            return (course, minutes_until, "upcoming")
        elif section_start.time() <= current_time <= section_end.time():
            minutes_remaining = (section_end - now).total_seconds() / 60
            return (course, minutes_remaining, "ongoing")
    for days_ahead in range(1, 8):
        future_weekday = ((current_weekday - 1 + days_ahead) % 7) + 1
        future_week = current_week
        if future_weekday < current_weekday:
            future_week += 1
        future_courses = [c for c in courses if c.weekday == future_weekday and future_week in c.weeks]
        if future_courses:
            next_course = min(future_courses, key=lambda c: c.start_section)
            return (next_course, days_ahead * 24 * 60, "future")
    return None


class CourseCard(BoxLayout):
    def __init__(self, course: Course, status: str = "", **kwargs):
        super().__init__(orientation='vertical', **kwargs)
        self.size_hint_y = None
        self.height = dp(120)
        self.padding = dp(15)
        self.spacing = dp(5)
        
        self.course = course
        self.status = status
        
        with self.canvas.before:
            if status == "ongoing":
                Color(0.12, 0.26, 0.20, 1)
            elif status == "upcoming":
                Color(0.29, 0.22, 0.15, 1)
            else:
                Color(0.18, 0.18, 0.18, 1)
            self.bg_rect = RoundedRectangle(pos=self.pos, size=self.size, radius=[dp(15)])
        
        self.bind(pos=self.update_rect, size=self.update_rect)
        
        status_icon = "🔴 正在上课" if status == "ongoing" else "🟡 即将上课" if status == "upcoming" else "📚"
        status_color = (0.31, 0.78, 0.47, 1) if status == "ongoing" else (1, 0.72, 0.30, 1) if status == "upcoming" else (0.67, 0.67, 0.67, 1)
        
        status_label = Label(
            text=status_icon,
            size_hint_y=None,
            height=dp(20),
            halign='left',
            valign='middle',
            color=status_color,
            font_size=dp(14)
        )
        status_label.bind(texture_size=status_label.setter('size'))
        self.add_widget(status_label)
        
        name_label = Label(
            text=course.name,
            size_hint_y=None,
            height=dp(30),
            halign='left',
            valign='middle',
            color=(1, 1, 1, 1),
            font_size=dp(18),
            bold=True
        )
        name_label.bind(texture_size=name_label.setter('size'))
        self.add_widget(name_label)
        
        start_time = SECTION_TIMES.get(course.start_section, (8, 0))
        end_time = SECTION_TIMES.get(course.end_section, (8, 0))
        time_str = f"{start_time[0]:02d}:{start_time[1]:02d} - {end_time[0]:02d}:{end_time[1]:02d}"
        
        info_text = f"⏰ {time_str}    📍 {course.location}    👨‍🏫 {course.teacher}"
        info_label = Label(
            text=info_text,
            size_hint_y=None,
            height=dp(25),
            halign='left',
            valign='middle',
            color=(0.7, 0.7, 0.7, 1),
            font_size=dp(12)
        )
        info_label.bind(texture_size=info_label.setter('size'))
        self.add_widget(info_label)
    
    def update_rect(self, instance, value):
        self.bg_rect.pos = instance.pos
        self.bg_rect.size = instance.size


class RoundedButton(Button):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.background_normal = ''
        self.background_color = (0.2, 0.4, 0.8, 1)
        self.color = (1, 1, 1, 1)
        self.font_size = dp(16)
        self.size_hint_y = None
        self.height = dp(50)
        
        with self.canvas.before:
            Color(0.2, 0.4, 0.8, 1)
            self.bg_rect = RoundedRectangle(pos=self.pos, size=self.size, radius=[dp(25)])
        
        self.bind(pos=self.update_rect, size=self.update_rect)
    
    def update_rect(self, instance, value):
        self.bg_rect.pos = instance.pos
        self.bg_rect.size = instance.size


class MainLayout(BoxLayout):
    def __init__(self, **kwargs):
        super().__init__(orientation='vertical', **kwargs)
        self.padding = dp(20)
        self.spacing = dp(15)
        
        self.courses = []
        self.semester_start = datetime(2026, 3, 2)
        self.reminded_courses = set()
        self.file_path = ""
        
        with self.canvas.before:
            Color(0.1, 0.1, 0.12, 1)
            self.bg_rect = RoundedRectangle(pos=self.pos, size=self.size)
        self.bind(pos=self.update_rect, size=self.update_rect)
        
        self.setup_ui()
        Clock.schedule_interval(self.update_ui, 1)
    
    def update_rect(self, instance, value):
        self.bg_rect.pos = instance.pos
        self.bg_rect.size = instance.size
    
    def setup_ui(self):
        header = BoxLayout(size_hint_y=None, height=dp(50))
        
        title = Label(
            text='🎓 课程提醒',
            halign='left',
            valign='middle',
            color=(1, 1, 1, 1),
            font_size=dp(26),
            bold=True,
            size_hint_x=0.7
        )
        title.bind(texture_size=title.setter('size'))
        
        self.theme_btn = RoundedButton(
            text='🌙',
            size_hint_x=0.3,
            size_hint_y=None,
            height=dp(40)
        )
        header.add_widget(title)
        header.add_widget(self.theme_btn)
        self.add_widget(header)
        
        info_card = BoxLayout(orientation='vertical', size_hint_y=None, height=dp(80), padding=dp(15))
        with info_card.canvas.before:
            Color(0.09, 0.14, 0.28, 1)
            info_bg = RoundedRectangle(pos=info_card.pos, size=info_card.size, radius=[dp(15)])
        info_card.bind(pos=lambda i, v: setattr(info_bg, 'pos', v), size=lambda i, v: setattr(info_bg, 'size', v))
        
        self.week_label = Label(
            text='第 ? 周',
            size_hint_y=None,
            height=dp(30),
            halign='center',
            valign='middle',
            color=(0.4, 0.7, 1, 1),
            font_size=dp(18),
            bold=True
        )
        self.date_label = Label(
            text='',
            size_hint_y=None,
            height=dp(25),
            halign='center',
            valign='middle',
            color=(0.5, 0.75, 1, 1),
            font_size=dp(14)
        )
        info_card.add_widget(self.week_label)
        info_card.add_widget(self.date_label)
        self.add_widget(info_card)
        
        countdown_card = BoxLayout(orientation='vertical', size_hint_y=None, height=dp(100), padding=dp(15))
        with countdown_card.canvas.before:
            Color(0.24, 0.15, 0.14, 1)
            countdown_bg = RoundedRectangle(pos=countdown_card.pos, size=countdown_card.size, radius=[dp(15)])
        countdown_card.bind(pos=lambda i, v: setattr(countdown_bg, 'pos', v), size=lambda i, v: setattr(countdown_bg, 'size', v))
        
        self.countdown_title = Label(
            text='⏱️ 距离下一节课',
            size_hint_y=None,
            height=dp(25),
            halign='center',
            valign='middle',
            color=(1, 0.72, 0.30, 1),
            font_size=dp(14)
        )
        self.countdown_label = Label(
            text='--:--:--',
            size_hint_y=None,
            height=dp(50),
            halign='center',
            valign='middle',
            color=(1, 0.6, 0.2, 1),
            font_size=dp(36),
            bold=True
        )
        countdown_card.add_widget(self.countdown_title)
        countdown_card.add_widget(self.countdown_label)
        self.add_widget(countdown_card)
        
        self.load_btn = RoundedButton(
            text='📂 选择课程表文件',
            size_hint_y=None,
            height=dp(50)
        )
        self.load_btn.bind(on_press=self.show_file_chooser)
        self.add_widget(self.load_btn)
        
        scroll = ScrollView()
        self.course_list = BoxLayout(orientation='vertical', spacing=dp(10), size_hint_y=None)
        self.course_list.bind(minimum_height=self.course_list.setter('height'))
        scroll.add_widget(self.course_list)
        self.add_widget(scroll)
        
        self.status_label = Label(
            text='请选择课程表文件',
            size_hint_y=None,
            height=dp(30),
            halign='center',
            valign='middle',
            color=(0.6, 0.6, 0.6, 1),
            font_size=dp(12)
        )
        self.add_widget(self.status_label)
    
    def show_file_chooser(self, instance):
        content = BoxLayout(orientation='vertical')
        
        file_chooser = FileChooserListView(
            path='/storage/emulated/0/' if os.path.exists('/storage/emulated/0/') else '.',
            filters=['*.xlsx']
        )
        content.add_widget(file_chooser)
        
        btn_layout = BoxLayout(size_hint_y=None, height=dp(50))
        cancel_btn = Button(text='取消')
        select_btn = Button(text='选择')
        btn_layout.add_widget(cancel_btn)
        btn_layout.add_widget(select_btn)
        content.add_widget(btn_layout)
        
        popup = Popup(
            title='选择课程表文件',
            content=content,
            size_hint=(0.9, 0.9)
        )
        
        cancel_btn.bind(on_press=popup.dismiss)
        select_btn.bind(on_press=lambda x: self.load_file(file_chooser.selection, popup))
        
        popup.open()
    
    def load_file(self, selection, popup):
        if selection:
            self.file_path = selection[0]
            try:
                self.courses = parse_xlsx_schedule(self.file_path)
                self.status_label.text = f'✅ 已加载 {len(self.courses)} 门课程'
            except Exception as e:
                self.status_label.text = f'❌ 加载失败: {str(e)}'
        popup.dismiss()
    
    def format_countdown(self, minutes: float) -> str:
        total_seconds = int(minutes * 60)
        hours = total_seconds // 3600
        minutes = (total_seconds % 3600) // 60
        seconds = total_seconds % 60
        if hours > 0:
            return f"{hours:02d}:{minutes:02d}:{seconds:02d}"
        return f"{minutes:02d}:{seconds:02d}"
    
    def update_ui(self, dt):
        now = datetime.now()
        current_week = get_current_week(self.semester_start)
        
        self.week_label.text = f'第 {current_week} 周 · {WEEKDAY_NAMES[now.weekday()]}'
        self.date_label.text = now.strftime('%Y年%m月%d日 %H:%M:%S')
        
        self.course_list.clear_widgets()
        
        if not self.courses:
            return
        
        result = get_next_course(self.courses, current_week)
        
        if result:
            course, minutes, status = result
            self.countdown_label.text = self.format_countdown(minutes)
            
            if status == "ongoing":
                self.countdown_title.text = '⏱️ 课程剩余时间'
            elif status == "upcoming":
                self.countdown_title.text = '⏱️ 距离上课还有'
                
                course_key = f"{course.code}_{now.strftime('%Y%m%d')}_{course.start_section}"
                if minutes <= 10 and course_key not in self.reminded_courses:
                    self.reminded_courses.add(course_key)
                    self.show_reminder_popup(course, minutes)
            else:
                self.countdown_title.text = '⏱️ 距离下一节课'
            
            card = CourseCard(course, status)
            self.course_list.add_widget(card)
    
    def show_reminder_popup(self, course: Course, minutes: float):
        content = BoxLayout(orientation='vertical', padding=dp(20), spacing=dp(10))
        
        content.add_widget(Label(
            text='🔔 即将上课!',
            font_size=dp(24),
            bold=True,
            color=(1, 0.6, 0.2, 1),
            size_hint_y=None,
            height=dp(40)
        ))
        
        content.add_widget(Label(
            text=f'📚 {course.name}',
            font_size=dp(18),
            bold=True,
            color=(1, 1, 1, 1),
            size_hint_y=None,
            height=dp(30)
        ))
        
        content.add_widget(Label(
            text=f'📍 {course.location}',
            font_size=dp(14),
            color=(0.7, 0.7, 0.7, 1),
            size_hint_y=None,
            height=dp(25)
        ))
        
        content.add_widget(Label(
            text=f'⏱️ 还有 {int(minutes)} 分钟',
            font_size=dp(16),
            color=(1, 0.72, 0.30, 1),
            size_hint_y=None,
            height=dp(30)
        ))
        
        ok_btn = RoundedButton(text='我知道了')
        content.add_widget(ok_btn)
        
        popup = Popup(
            title='',
            content=content,
            size_hint=(0.85, 0.45),
            background_color=(0.15, 0.15, 0.18, 1),
            separator_color=(0.2, 0.2, 0.25, 1)
        )
        
        ok_btn.bind(on_press=popup.dismiss)
        popup.open()


class ClassReminderApp(App):
    def build(self):
        self.title = '课程提醒'
        return MainLayout()


if __name__ == '__main__':
    ClassReminderApp().run()
